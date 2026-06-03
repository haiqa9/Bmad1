import openpyxl
from collections import Counter

wb = openpyxl.load_workbook('D:/Bmad/ITAM/List of IT Assets.xlsx')
ws = wb['Laptop Record ']

GREEN = 'FF00FF00'
rows = []
for row_idx in range(2, ws.max_row + 1):
    cell_a = ws.cell(row=row_idx, column=1)
    is_green = cell_a.fill.fgColor.rgb == GREEN if cell_a.fill and cell_a.fill.fgColor else False
    
    row_data = []
    for col_idx in range(1, 19):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val is None:
            row_data.append('')
        elif hasattr(val, 'isoformat'):
            row_data.append(val.isoformat())
        elif isinstance(val, bool):
            row_data.append('Yes' if val else 'No')
        else:
            row_data.append(str(val).strip())
    
    if any(row_data):
        rows.append({
            'row': row_idx,
            'green': is_green,
            'status': row_data[3],
            'employee': row_data[1],
            'serial': row_data[5],
            'data': row_data
        })

print(f'Total rows: {len(rows)}')
green_rows = [r for r in rows if r['green']]
print(f'Green (Issued) rows: {len(green_rows)}')
print(f'Non-green rows: {len(rows) - len(green_rows)}')

print('\n=== Green rows sample ===')
for r in green_rows[:5]:
    print(f'Row {r["row"]}: {r["employee"]} | Serial: {r["serial"]} | Status: {r["status"]}')

print('\n=== Non-green rows sample ===')
non_green = [r for r in rows if not r['green']]
for r in non_green[:5]:
    print(f'Row {r["row"]}: {r["employee"]} | Serial: {r["serial"]} | Status: {r["status"]}')

# Check for duplicates by serial number
serials = [r['serial'] for r in rows if r['serial']]
dupes = {s: c for s, c in Counter(serials).items() if c > 1}
print(f'\nDuplicate serials: {len(dupes)}')
for s, c in list(dupes.items())[:10]:
    print(f'  {s}: {c} times')

# Also check duplicates by employee + serial
emp_serial = [(r['employee'], r['serial']) for r in rows if r['employee'] and r['serial']]
dupes2 = {k: c for k, c in Counter(emp_serial).items() if c > 1}
print(f'\nDuplicate (employee, serial): {len(dupes2)}')

# Let's see the duplicates for the first one
if dupes:
    first_serial = list(dupes.keys())[0]
    print(f'\n=== All entries for serial "{first_serial}" ===')
    for r in rows:
        if r['serial'] == first_serial:
            print(f'Row {r["row"]}: Green={r["green"]} | Employee={r["employee"]} | Status={r["status"]}')
