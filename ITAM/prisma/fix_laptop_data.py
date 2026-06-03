import openpyxl
import json
from collections import defaultdict
from datetime import datetime

def to_str(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, bool):
        return "Yes" if v else "No"
    return str(v).strip()

wb = openpyxl.load_workbook('D:/Bmad/ITAM/List of IT Assets.xlsx')
ws = wb['Laptop Record ']

GREEN = 'FF00FF00'

# Read all rows with green highlight info
raw_rows = []
for row_idx in range(2, ws.max_row + 1):
    cell_a = ws.cell(row=row_idx, column=1)
    is_green = False
    if cell_a.fill and cell_a.fill.fgColor and cell_a.fill.fgColor.rgb:
        is_green = cell_a.fill.fgColor.rgb == GREEN
    
    row_data = []
    for col_idx in range(1, 19):
        val = ws.cell(row=row_idx, column=col_idx).value
        row_data.append(to_str(val))
    
    if any(row_data):
        raw_rows.append({
            'row': row_idx,
            'green': is_green,
            'data': row_data,
            'serial': row_data[5].strip(),
            'employee': row_data[1].strip(),
            'status': row_data[3].strip(),
        })

print(f'Total raw rows: {len(raw_rows)}')
print(f'Green (Issued) rows: {sum(1 for r in raw_rows if r["green"])}')
print(f'Non-green rows: {sum(1 for r in raw_rows if not r["green"])}')

# Deduplicate by serial number
serial_groups = defaultdict(list)
no_serial_rows = []

for r in raw_rows:
    if r['serial'] and r['serial'] != '-':
        serial_groups[r['serial']].append(r)
    else:
        no_serial_rows.append(r)

deduped = []
for serial, group in serial_groups.items():
    green_in_group = [g for g in group if g['green']]
    if green_in_group:
        for g in green_in_group:
            deduped.append(g)
    else:
        deduped.append(group[-1])

deduped.extend(no_serial_rows)

print(f'\nAfter deduplication: {len(deduped)}')
print(f'  Green (Issued): {sum(1 for r in deduped if r["green"])}')
print(f'  Non-green (Returned): {sum(1 for r in deduped if not r["green"])}')

# Prepare final rows with sequential sr numbers (1, 2, 3...)
final_rows = []
for idx, r in enumerate(deduped, start=1):
    # Replace saq (index 0) with sequential number, append isIssued
    row = [str(idx)] + r['data'][1:] + ['Yes' if r['green'] else 'No']
    final_rows.append(row)

# Update seed-data.json
with open('D:/Bmad/ITAM/prisma/seed-data.json', 'r', encoding='utf-8') as f:
    seed_data = json.load(f)

# Update headers: saq -> sr, add isIssued
headers = seed_data['laptop-record']['headers']
if headers[0] == 'saq':
    headers[0] = 'sr'
if headers[-1] != 'isIssued':
    headers.append('isIssued')

seed_data['laptop-record']['headers'] = headers
seed_data['laptop-record']['rows'] = final_rows

with open('D:/Bmad/ITAM/prisma/seed-data.json', 'w', encoding='utf-8') as f:
    json.dump(seed_data, f, ensure_ascii=False, indent=2)

print('\nUpdated seed-data.json with sr=1..{}'.format(len(final_rows)))

# Show duplicate examples
print('\n=== Removed duplicates sample ===')
count = 0
for serial, group in serial_groups.items():
    if len(group) > 1 and count < 5:
        count += 1
        print(f'\nSerial {serial} ({len(group)} entries -> kept green or last):')
        for g in group:
            marker = ' [KEPT]' if g in deduped else ' [REMOVED]'
            print(f'  Row {g["row"]}: Green={g["green"]} | {g["employee"]} | {g["status"]}{marker}')
