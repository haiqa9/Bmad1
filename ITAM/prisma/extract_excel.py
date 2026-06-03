import openpyxl
import json
from datetime import datetime

def to_str(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, bool):
        return "Yes" if v else "No"
    return str(v).strip()

def get_headers(ws, row_num):
    """Get non-empty headers from a specific row, stopping at the last non-empty cell."""
    row = list(ws[row_num])
    values = []
    last_non_empty = -1
    for i, cell in enumerate(row):
        val = to_str(cell.value)
        values.append(val)
        if val:
            last_non_empty = i
    if last_non_empty >= 0:
        values = values[:last_non_empty + 1]
    return values

def extract_sheet(ws, header_row, data_start_row):
    headers = get_headers(ws, header_row)
    rows = []
    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        row_vals = [to_str(v) for v in row[:len(headers)]]
        if any(v for v in row_vals):
            rows.append(row_vals)
    return headers, rows

wb = openpyxl.load_workbook('D:/Bmad/ITAM/List of IT Assets.xlsx', data_only=True)

sheets = {}

# 1. Laptop Record — Row 1 headers, data from row 2
h, r = extract_sheet(wb['Laptop Record '], 1, 2)
sheets['laptop-record'] = {"name": "Laptop Record", "headers": h, "rows": r}

# 2. Servers - Devices — Row 2 headers, data from row 3
h, r = extract_sheet(wb['Servers - Devices'], 2, 3)
sheets['servers-devices'] = {"name": "Servers - Devices", "headers": h, "rows": r}

# 3. iloidrac — Row 1 headers, data from row 2
h, r = extract_sheet(wb['iloidrac'], 1, 2)
sheets['iloidrac'] = {"name": "iloidrac", "headers": h, "rows": r}

# 4. Cloud VM List — Row 1 headers, data from row 2
h, r = extract_sheet(wb['Cloud VM List'], 1, 2)
sheets['cloud-vm-list'] = {"name": "Cloud VM List", "headers": h, "rows": r}

# 5. Lab VM List — Row 1 headers, data from row 2
h, r = extract_sheet(wb['Lab VM List'], 1, 2)
sheets['lab-vm-list'] = {"name": "Lab VM List", "headers": h, "rows": r}

# 6. public FQDN — Row 1 headers, data from row 2
h, r = extract_sheet(wb['public FQDN'], 1, 2)
sheets['public-fqdn'] = {"name": "public FQDN", "headers": h, "rows": r}

# 7. GatePass — Row 3 headers, data from row 4
h, r = extract_sheet(wb['GatePass'], 3, 4)
sheets['gatepass'] = {"name": "GatePass", "headers": h, "rows": r}

# 8. Received Items — Row 2 headers, data from row 3
h, r = extract_sheet(wb['Received Items'], 2, 3)
sheets['received-items'] = {"name": "Received Items", "headers": h, "rows": r}

# 9. Ports-Detail — Row 1 headers, data from row 3 (row 2 is sub-header)
h, r = extract_sheet(wb['Ports-Detail'], 1, 3)
sheets['ports-detail'] = {"name": "Ports-Detail", "headers": h, "rows": r}

# 10. Free VMs — Row 1 headers, data from row 2
h, r = extract_sheet(wb['Free VMs'], 1, 2)
sheets['free-vms'] = {"name": "Free VMs", "headers": h, "rows": r}

# 11. Sheet38 — Row 3 headers, data from row 4
h, r = extract_sheet(wb['Sheet38'], 3, 4)
sheets['sheet38'] = {"name": "Sheet38", "headers": h, "rows": r}

# Summary
for key, val in sheets.items():
    print(f"{key}: {len(val['headers'])} cols, {len(val['rows'])} rows")
    print(f"  Headers: {val['headers']}")

with open('D:/Bmad/ITAM/prisma/seed-data.json', 'w', encoding='utf-8') as f:
    json.dump(sheets, f, ensure_ascii=False, indent=2)
print("\nWritten to prisma/seed-data.json")
